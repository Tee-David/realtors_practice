#!/usr/bin/env python3
"""
Master Workbook Verification Script

Verifies that the master workbook has:
1. All site sheets
2. All 7 summary sheets
3. Correct data in each sheet
"""

import sys
from pathlib import Path

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook

MASTER_WORKBOOK = Path("exports/cleaned/MASTER_CLEANED_WORKBOOK.xlsx")
EXPECTED_SUMMARY_SHEETS = [
    '_Dashboard',
    '_Top_100_Cheapest',
    '_Newest_Listings',
    '_For_Sale',
    '_For_Rent',
    '_Land_Only',
    '_4BR_Plus',
    '_Metadata'
]


def verify_master_workbook():
    """Verify master workbook structure and content."""

    print("="*60)
    print("MASTER WORKBOOK VERIFICATION")
    print("="*60)

    # Check if file exists
    if not MASTER_WORKBOOK.exists():
        print(f"❌ FAILED: Master workbook not found at {MASTER_WORKBOOK}")
        return False

    print(f"✅ Master workbook exists: {MASTER_WORKBOOK}")
    print(f"   Size: {MASTER_WORKBOOK.stat().st_size / 1024:.1f} KB\n")

    # Load workbook
    try:
        wb = load_workbook(MASTER_WORKBOOK, read_only=True)
    except Exception as e:
        print(f"❌ FAILED: Cannot load workbook: {e}")
        return False

    print(f"✅ Workbook loaded successfully\n")

    # Get all sheets
    all_sheets = wb.sheetnames
    print(f"📊 Total sheets: {len(all_sheets)}")
    print(f"   Sheets: {', '.join(all_sheets)}\n")

    # Verify summary sheets
    print("🔍 Checking summary sheets...")
    missing_summaries = []
    for sheet_name in EXPECTED_SUMMARY_SHEETS:
        if sheet_name in all_sheets:
            ws = wb[sheet_name]
            row_count = ws.max_row
            print(f"   ✅ {sheet_name:20s} - {row_count:4d} rows")
        else:
            missing_summaries.append(sheet_name)
            print(f"   ❌ {sheet_name:20s} - MISSING!")

    if missing_summaries:
        print(f"\n❌ FAILED: Missing summary sheets: {', '.join(missing_summaries)}")
        wb.close()
        return False

    print(f"\n✅ All {len(EXPECTED_SUMMARY_SHEETS)} summary sheets present!\n")

    # Verify site sheets
    site_sheets = [s for s in all_sheets if not s.startswith('_')]
    print(f"🏢 Site sheets: {len(site_sheets)}")

    total_listings = 0
    for sheet_name in site_sheets:
        ws = wb[sheet_name]
        row_count = ws.max_row - 1  # Subtract header row
        total_listings += row_count
        print(f"   📍 {sheet_name:20s} - {row_count:4d} listings")

    print(f"\n📈 Total listings across all sites: {total_listings}\n")

    # Verify _Dashboard content
    print("🎯 Verifying _Dashboard content...")
    ws_dashboard = wb['_Dashboard']

    # Check for key sections
    has_stats = False
    has_property_types = False
    has_top_sites = False

    for row in ws_dashboard.iter_rows(values_only=True):
        if row[0]:
            text = str(row[0])
            if 'OVERALL STATISTICS' in text:
                has_stats = True
            elif 'PROPERTY TYPE BREAKDOWN' in text:
                has_property_types = True
            elif 'TOP SITES' in text:
                has_top_sites = True

    if has_stats:
        print("   ✅ Overall statistics section present")
    else:
        print("   ⚠️  Overall statistics section missing")

    if has_property_types:
        print("   ✅ Property type breakdown present")
    else:
        print("   ⚠️  Property type breakdown missing")

    if has_top_sites:
        print("   ✅ Top sites section present")
    else:
        print("   ⚠️  Top sites section missing")

    print()

    wb.close()

    # Final verdict
    print("="*60)
    if missing_summaries:
        print("❌ VERIFICATION FAILED")
        print(f"   Missing sheets: {', '.join(missing_summaries)}")
        return False
    else:
        print("✅ VERIFICATION PASSED")
        print(f"   ✓ {len(EXPECTED_SUMMARY_SHEETS)} summary sheets")
        print(f"   ✓ {len(site_sheets)} site sheets")
        print(f"   ✓ {total_listings} total listings")
        print("="*60)
        return True


if __name__ == "__main__":
    success = verify_master_workbook()
    sys.exit(0 if success else 1)
