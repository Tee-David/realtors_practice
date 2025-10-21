# core/master_workbook.py
"""
Master Workbook Management Module

Handles:
- Creating and managing MASTER_CLEANED_WORKBOOK.xlsx
- Per-site sheet management (create, append)
- Append-only idempotent logic
- Workbook optimization (freeze panes, filters, column widths)
- Per-site CSV and Parquet exports
"""

import logging
import json
from pathlib import Path
from typing import Dict, List, Set, Optional, Any
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

from core.data_cleaner import CANONICAL_SCHEMA

# Metadata file for tracking master workbook state
METADATA_FILE = Path("exports/cleaned/metadata.json")


class MasterWorkbookManager:
    """Manages the master cleaned workbook with per-site sheets."""

    def __init__(self, workbook_path: Path):
        self.workbook_path = workbook_path
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        self.metadata = self._load_metadata()

    def _load_metadata(self) -> Dict:
        """Load metadata tracking master workbook state."""
        if METADATA_FILE.exists():
            try:
                with open(METADATA_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                logging.warning(f"Failed to load metadata: {e}")
                return {}
        return {}

    def _save_metadata(self):
        """Save metadata to disk."""
        try:
            METADATA_FILE.parent.mkdir(parents=True, exist_ok=True)
            with open(METADATA_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.metadata, f, indent=2)
        except Exception as e:
            logging.error(f"Failed to save metadata: {e}")

    def _ensure_workbook_exists(self):
        """Ensure master workbook file exists."""
        if not self.workbook_path.exists():
            logging.info(f"Creating new master workbook: {self.workbook_path}")
            wb = Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Create metadata sheet
            ws_meta = wb.create_sheet('_Metadata')
            ws_meta.append(['Field', 'Value'])
            ws_meta.append(['Created', datetime.now().isoformat()])
            ws_meta.append(['Total Sites', 0])
            ws_meta.append(['Total Records', 0])
            ws_meta.append(['Last Updated', datetime.now().isoformat()])

            # Format metadata sheet
            ws_meta['A1'].font = Font(bold=True)
            ws_meta['B1'].font = Font(bold=True)
            ws_meta.column_dimensions['A'].width = 20
            ws_meta.column_dimensions['B'].width = 40

            wb.save(self.workbook_path)
            logging.info("Master workbook created successfully")

    def _get_existing_hashes(self, sheet_name: str) -> Set[str]:
        """
        Get set of existing record hashes from a sheet.

        Used for deduplication during append.
        """
        try:
            wb = load_workbook(self.workbook_path, read_only=True)

            if sheet_name not in wb.sheetnames:
                return set()

            ws = wb[sheet_name]

            # Find hash column index
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            try:
                hash_col_idx = header_row.index('hash') + 1  # 1-indexed
            except ValueError:
                logging.warning(f"No 'hash' column found in {sheet_name}")
                return set()

            # Collect all hashes
            hashes = set()
            for row in ws.iter_rows(min_row=2, min_col=hash_col_idx, max_col=hash_col_idx, values_only=True):
                if row[0]:
                    hashes.add(str(row[0]))

            wb.close()
            return hashes

        except Exception as e:
            logging.error(f"Error reading existing hashes from {sheet_name}: {e}")
            return set()

    def _create_site_sheet(self, site_key: str) -> None:
        """
        Create a new sheet for a site.

        Sets up headers, formatting, and filters.
        """
        try:
            wb = load_workbook(self.workbook_path)

            # Create sheet
            ws = wb.create_sheet(site_key)

            # Add headers
            ws.append(CANONICAL_SCHEMA)

            # Format header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Set column widths
            column_widths = {
                'A': 50,  # title
                'B': 15,  # price
                'C': 15,  # price_per_sqm
                'D': 15,  # price_per_bedroom
                'E': 30,  # location
                'F': 30,  # estate_name
                'G': 20,  # property_type
                'H': 10,  # bedrooms
                'I': 10,  # bathrooms
                'J': 10,  # toilets
                'K': 10,  # bq
                'L': 15,  # land_size
                'M': 15,  # title_tag
                'N': 50,  # description
                'O': 20,  # promo_tags
                'P': 15,  # initial_deposit
                'Q': 20,  # payment_plan
                'R': 15,  # service_charge
                'S': 20,  # launch_timeline
                'T': 30,  # agent_name
                'U': 30,  # contact_info
                'V': 50,  # images
                'W': 60,  # listing_url
                'X': 30,  # source
                'Y': 20,  # scrape_timestamp
                'Z': 25,  # coordinates
                'AA': 70, # hash
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Add auto-filter (handle columns beyond Z)
            last_col_idx = len(CANONICAL_SCHEMA)
            if last_col_idx <= 26:
                last_col = chr(64 + last_col_idx)
            else:
                # For columns beyond Z (AA, AB, etc.)
                last_col = chr(64 + (last_col_idx - 1) // 26) + chr(65 + (last_col_idx - 1) % 26)
            ws.auto_filter.ref = f"A1:{last_col}1"

            wb.save(self.workbook_path)
            logging.info(f"Created new sheet: {site_key}")

        except Exception as e:
            logging.error(f"Failed to create sheet {site_key}: {e}")
            raise

    def _append_records_to_sheet(self, site_key: str, new_records: List[Dict]) -> int:
        """
        Append new records to a site sheet.

        Returns number of records actually appended (after deduplication).
        """
        try:
            # Get existing hashes for deduplication
            existing_hashes = self._get_existing_hashes(site_key)
            logging.debug(f"Sheet {site_key} has {len(existing_hashes)} existing records")

            # Filter out duplicates
            records_to_append = [
                r for r in new_records
                if r.get('hash') not in existing_hashes
            ]

            if not records_to_append:
                logging.info(f"No new records to append to {site_key} (all duplicates)")
                return 0

            # Load workbook
            wb = load_workbook(self.workbook_path)
            ws = wb[site_key]

            # Append records
            for record in records_to_append:
                row = [record.get(field, '') for field in CANONICAL_SCHEMA]
                ws.append(row)

            wb.save(self.workbook_path)
            logging.info(f"Appended {len(records_to_append)} new records to {site_key}")

            return len(records_to_append)

        except Exception as e:
            logging.error(f"Failed to append records to {site_key}: {e}")
            raise

    def _update_metadata_sheet(self):
        """Update _Metadata sheet with current statistics."""
        try:
            wb = load_workbook(self.workbook_path)

            # Count sites and records
            site_count = len([s for s in wb.sheetnames if s != '_Metadata'])
            total_records = 0

            for sheet_name in wb.sheetnames:
                if sheet_name == '_Metadata':
                    continue
                ws = wb[sheet_name]
                total_records += ws.max_row - 1  # Subtract header row

            # Update metadata sheet
            ws_meta = wb['_Metadata']
            ws_meta['B3'] = site_count
            ws_meta['B4'] = total_records
            ws_meta['B5'] = datetime.now().isoformat()

            wb.save(self.workbook_path)
            logging.debug("Updated metadata sheet")

        except Exception as e:
            logging.error(f"Failed to update metadata sheet: {e}")

    def append_to_site(self, site_key: str, records: List[Dict]) -> int:
        """
        Append records to a site sheet (idempotent).

        Creates sheet if it doesn't exist.
        Filters out duplicates based on hash.

        Returns number of new records appended.
        """
        if not records:
            logging.warning(f"No records to append for {site_key}")
            return 0

        # Ensure workbook exists
        self._ensure_workbook_exists()

        # Check if sheet exists
        wb = load_workbook(self.workbook_path, read_only=True)
        sheet_exists = site_key in wb.sheetnames
        wb.close()

        # Create sheet if needed
        if not sheet_exists:
            self._create_site_sheet(site_key)

        # Append records
        new_count = self._append_records_to_sheet(site_key, records)

        # Update metadata and summary sheets
        if new_count > 0:
            self._update_metadata_sheet()
            self._update_site_metadata(site_key, new_count)

            # Regenerate summary sheets with new data
            try:
                self._regenerate_summary_sheets()
            except Exception as e:
                logging.error(f"Failed to regenerate summary sheets: {e}")

        return new_count

    def _update_site_metadata(self, site_key: str, new_count: int):
        """Update metadata tracking for a specific site."""
        if site_key not in self.metadata:
            self.metadata[site_key] = {
                'total_records': 0,
                'first_added': datetime.now().isoformat(),
                'source_files': []
            }

        self.metadata[site_key]['total_records'] += new_count
        self.metadata[site_key]['last_updated'] = datetime.now().isoformat()

        self._save_metadata()

    def _load_all_data(self) -> pd.DataFrame:
        """
        Load all site data into a single DataFrame.

        Returns consolidated data from all non-summary sheets.
        """
        wb = load_workbook(self.workbook_path, read_only=True)
        all_data = []

        for sheet_name in wb.sheetnames:
            # Skip metadata and summary sheets
            if sheet_name.startswith('_'):
                continue

            ws = wb[sheet_name]
            data = list(ws.values)

            if len(data) < 2:  # No data rows
                continue

            # Convert to dataframe
            cols = data[0]
            rows = data[1:]
            df = pd.DataFrame(rows, columns=cols)

            # Add source site
            df['_site'] = sheet_name

            all_data.append(df)

        wb.close()

        if not all_data:
            return pd.DataFrame()

        # Combine all data
        combined = pd.concat(all_data, ignore_index=True)

        # Convert numeric columns
        numeric_cols = ['price', 'price_per_sqm', 'price_per_bedroom', 'bedrooms',
                       'bathrooms', 'toilets', 'bq', 'land_size']
        for col in numeric_cols:
            if col in combined.columns:
                combined[col] = pd.to_numeric(combined[col], errors='coerce')

        # Parse timestamps
        if 'scrape_timestamp' in combined.columns:
            combined['scrape_timestamp'] = pd.to_datetime(combined['scrape_timestamp'], errors='coerce')

        return combined

    def _regenerate_summary_sheets(self):
        """Regenerate all summary sheets with latest data."""
        logging.info("Regenerating summary sheets...")

        # Load all data
        df = self._load_all_data()

        if df.empty:
            logging.warning("No data to generate summary sheets")
            return

        # Open workbook for writing
        wb = load_workbook(self.workbook_path)

        # Delete old summary sheets
        summary_sheets = ['_Dashboard', '_Top_100_Cheapest', '_Newest_Listings',
                         '_For_Sale', '_For_Rent', '_Land_Only', '_4BR_Plus']
        for sheet_name in summary_sheets:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

        # Generate each summary sheet
        self._create_dashboard_sheet(wb, df)
        self._create_top_cheapest_sheet(wb, df)
        self._create_newest_listings_sheet(wb, df)
        self._create_for_sale_sheet(wb, df)
        self._create_for_rent_sheet(wb, df)
        self._create_land_only_sheet(wb, df)
        self._create_4br_plus_sheet(wb, df)

        # Save workbook
        wb.save(self.workbook_path)
        logging.info("Summary sheets regenerated successfully")

    def _create_dashboard_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create dashboard summary sheet."""
        ws = wb.create_sheet('_Dashboard', 0)  # Insert as first sheet

        # Title
        ws['A1'] = 'LAGOS PROPERTY SCRAPER - DASHBOARD'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        row = 3

        # Overall statistics
        ws[f'A{row}'] = 'OVERALL STATISTICS'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        total_listings = len(df)
        total_sites = df['_site'].nunique() if '_site' in df.columns else 0
        avg_price = df['price'].mean() if 'price' in df.columns else 0

        ws[f'A{row}'] = 'Total Listings:'
        ws[f'B{row}'] = total_listings
        row += 1

        ws[f'A{row}'] = 'Total Sites:'
        ws[f'B{row}'] = total_sites
        row += 1

        ws[f'A{row}'] = 'Average Price:'
        ws[f'B{row}'] = f"₦{avg_price:,.0f}" if avg_price > 0 else "N/A"
        row += 2

        # Property type breakdown
        ws[f'A{row}'] = 'PROPERTY TYPE BREAKDOWN'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        if 'property_type' in df.columns:
            type_counts = df['property_type'].value_counts().head(10)
            for ptype, count in type_counts.items():
                ws[f'A{row}'] = str(ptype) if ptype else 'Unknown'
                ws[f'B{row}'] = int(count)
                row += 1

        row += 1

        # Top sites by listing count
        ws[f'A{row}'] = 'TOP SITES BY LISTING COUNT'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        if '_site' in df.columns:
            site_counts = df['_site'].value_counts().head(10)
            for site, count in site_counts.items():
                ws[f'A{row}'] = str(site)
                ws[f'B{row}'] = int(count)
                row += 1

        # Format columns
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def _create_top_cheapest_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create top 100 cheapest properties sheet."""
        ws = wb.create_sheet('_Top_100_Cheapest')

        # Filter valid prices and sort
        df_filtered = df[df['price'] > 0].copy() if 'price' in df.columns else df.copy()
        df_sorted = df_filtered.nsmallest(100, 'price') if 'price' in df_filtered.columns else df_filtered.head(100)

        # Select columns
        display_cols = ['title', 'price', 'bedrooms', 'location', 'property_type',
                       '_site', 'listing_url', 'scrape_timestamp']
        df_export = df_sorted[[c for c in display_cols if c in df_sorted.columns]]

        # Write to sheet
        self._write_dataframe_to_sheet(ws, df_export, 'Top 100 Cheapest Properties')

    def _create_newest_listings_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create newest listings sheet."""
        ws = wb.create_sheet('_Newest_Listings')

        # Sort by timestamp
        if 'scrape_timestamp' in df.columns:
            df_sorted = df.sort_values('scrape_timestamp', ascending=False).head(100)
        else:
            df_sorted = df.head(100)

        # Select columns
        display_cols = ['title', 'price', 'bedrooms', 'location', 'property_type',
                       '_site', 'listing_url', 'scrape_timestamp']
        df_export = df_sorted[[c for c in display_cols if c in df_sorted.columns]]

        self._write_dataframe_to_sheet(ws, df_export, 'Newest 100 Listings')

    def _create_for_sale_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create for-sale properties sheet."""
        ws = wb.create_sheet('_For_Sale')

        # Filter for sale properties
        mask = pd.Series(False, index=df.index)

        if 'title' in df.columns:
            mask |= df['title'].str.contains('for sale|sale', case=False, na=False)

        if 'price' in df.columns:
            mask |= (df['price'] > 10000000)  # Likely sale if > 10M

        df_filtered = df[mask].copy()

        # Sort by price
        if 'price' in df_filtered.columns and not df_filtered.empty:
            df_sorted = df_filtered.sort_values('price')
        else:
            df_sorted = df_filtered

        # Select columns
        display_cols = ['title', 'price', 'bedrooms', 'location', 'property_type',
                       '_site', 'listing_url']
        df_export = df_sorted[[c for c in display_cols if c in df_sorted.columns]]

        self._write_dataframe_to_sheet(ws, df_export, 'Properties For Sale')

    def _create_for_rent_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create for-rent properties sheet."""
        ws = wb.create_sheet('_For_Rent')

        # Filter for rent properties
        mask = pd.Series(False, index=df.index)

        if 'title' in df.columns:
            mask |= df['title'].str.contains('rent|shortlet|lease', case=False, na=False)

        if 'price' in df.columns:
            mask |= (df['price'] > 0) & (df['price'] < 1000000)  # Likely rent if < 1M

        df_filtered = df[mask].copy()

        # Sort by price
        if 'price' in df_filtered.columns and not df_filtered.empty:
            df_sorted = df_filtered.sort_values('price')
        else:
            df_sorted = df_filtered

        # Select columns
        display_cols = ['title', 'price', 'bedrooms', 'location', 'property_type',
                       '_site', 'listing_url']
        df_export = df_sorted[[c for c in display_cols if c in df_sorted.columns]]

        self._write_dataframe_to_sheet(ws, df_export, 'Properties For Rent')

    def _create_land_only_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create land-only properties sheet."""
        ws = wb.create_sheet('_Land_Only')

        # Filter for land
        mask = pd.Series(False, index=df.index)

        if 'property_type' in df.columns:
            mask |= df['property_type'].str.contains('land', case=False, na=False)

        if 'title' in df.columns:
            mask |= df['title'].str.contains('land|plot|acre|hectare', case=False, na=False)

        df_filtered = df[mask].copy()

        # Sort by price
        if 'price' in df_filtered.columns and not df_filtered.empty:
            df_sorted = df_filtered.sort_values('price')
        else:
            df_sorted = df_filtered

        # Select columns
        display_cols = ['title', 'price', 'land_size', 'location', 'title_tag',
                       '_site', 'listing_url']
        df_export = df_sorted[[c for c in display_cols if c in df_sorted.columns]]

        self._write_dataframe_to_sheet(ws, df_export, 'Land Only')

    def _create_4br_plus_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create 4+ bedroom properties sheet."""
        ws = wb.create_sheet('_4BR_Plus')

        # Filter for 4+ bedrooms
        if 'bedrooms' in df.columns:
            df_filtered = df[df['bedrooms'] >= 4].copy()
        else:
            df_filtered = df.head(0)  # Empty

        # Sort by price
        if 'price' in df_filtered.columns and not df_filtered.empty:
            df_sorted = df_filtered.sort_values('price')
        else:
            df_sorted = df_filtered

        # Select columns
        display_cols = ['title', 'price', 'bedrooms', 'bathrooms', 'location',
                       'property_type', '_site', 'listing_url']
        df_export = df_sorted[[c for c in display_cols if c in df_sorted.columns]]

        self._write_dataframe_to_sheet(ws, df_export, '4+ Bedroom Properties')

    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame, title: str):
        """Helper to write dataframe to worksheet with formatting."""
        # Title row
        ws['A1'] = title
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells(f'A1:{chr(65 + len(df.columns) - 1)}1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Write headers (row 2)
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=str(col_name))
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data (starting row 3)
        for row_idx, row_data in enumerate(df.itertuples(index=False), 3):
            for col_idx, value in enumerate(row_data, 1):
                # Format value
                if pd.isna(value):
                    formatted_value = ''
                elif isinstance(value, (int, float)):
                    formatted_value = value
                elif isinstance(value, datetime):
                    formatted_value = value.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    formatted_value = str(value)

                ws.cell(row=row_idx, column=col_idx, value=formatted_value)

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(df.columns, 1):
            col_letter = chr(64 + col_idx) if col_idx <= 26 else f"{chr(64 + (col_idx - 1) // 26)}{chr(65 + (col_idx - 1) % 26)}"

            # Set reasonable widths based on column name
            if 'title' in str(col_name).lower() or 'description' in str(col_name).lower():
                ws.column_dimensions[col_letter].width = 50
            elif 'url' in str(col_name).lower():
                ws.column_dimensions[col_letter].width = 60
            elif 'location' in str(col_name).lower() or 'estate' in str(col_name).lower():
                ws.column_dimensions[col_letter].width = 30
            elif 'price' in str(col_name).lower():
                ws.column_dimensions[col_letter].width = 15
            else:
                ws.column_dimensions[col_letter].width = 20

        # Freeze header rows
        ws.freeze_panes = 'A3'

        # Add auto-filter
        last_col = chr(64 + len(df.columns)) if len(df.columns) <= 26 else f"{chr(64 + (len(df.columns) - 1) // 26)}{chr(65 + (len(df.columns) - 1) % 26)}"
        ws.auto_filter.ref = f"A2:{last_col}2"

    def export_site_to_csv(self, site_key: str, output_dir: Path):
        """
        Export a site sheet to CSV file.

        Creates: exports/cleaned/<site>/<site>_cleaned.csv
        """
        try:
            output_dir.mkdir(parents=True, exist_ok=True)
            csv_path = output_dir / f"{site_key}_cleaned.csv"

            # Read sheet
            wb = load_workbook(self.workbook_path, read_only=True)
            if site_key not in wb.sheetnames:
                logging.warning(f"Sheet {site_key} not found")
                return

            ws = wb[site_key]

            # Convert to dataframe
            data = ws.values
            cols = next(data)
            df = pd.DataFrame(data, columns=cols)

            # Export
            df.to_csv(csv_path, index=False, encoding='utf-8-sig')
            logging.info(f"Exported {site_key} to {csv_path}")

            wb.close()

        except Exception as e:
            logging.error(f"Failed to export {site_key} to CSV: {e}")

    def export_site_to_parquet(self, site_key: str, output_dir: Path):
        """
        Export a site sheet to Parquet file.

        Creates: exports/cleaned/<site>/<site>_cleaned.parquet
        """
        try:
            output_dir.mkdir(parents=True, exist_ok=True)
            parquet_path = output_dir / f"{site_key}_cleaned.parquet"

            # Read sheet
            wb = load_workbook(self.workbook_path, read_only=True)
            if site_key not in wb.sheetnames:
                logging.warning(f"Sheet {site_key} not found")
                return

            ws = wb[site_key]

            # Convert to dataframe
            data = ws.values
            cols = next(data)
            df = pd.DataFrame(data, columns=cols)

            # Export
            df.to_parquet(parquet_path, index=False, engine='pyarrow')
            logging.info(f"Exported {site_key} to {parquet_path}")

            wb.close()

        except Exception as e:
            logging.error(f"Failed to export {site_key} to Parquet: {e}")


def append_to_master(site_key: str, records: List[Dict], workbook_path: Path) -> int:
    """
    Main entry point: append records to master workbook for a site.

    Args:
        site_key: Site identifier
        records: List of cleaned, normalized records
        workbook_path: Path to MASTER_CLEANED_WORKBOOK.xlsx

    Returns:
        Number of new records appended
    """
    manager = MasterWorkbookManager(workbook_path)
    new_count = manager.append_to_site(site_key, records)

    # Also export to CSV and Parquet
    if new_count > 0:
        output_dir = workbook_path.parent / site_key
        try:
            manager.export_site_to_csv(site_key, output_dir)
        except Exception as e:
            logging.error(f"CSV export failed for {site_key}: {e}")

        try:
            manager.export_site_to_parquet(site_key, output_dir)
        except Exception as e:
            # Parquet may fail if pyarrow not installed
            logging.warning(f"Parquet export skipped for {site_key}: {e}")

    return new_count
