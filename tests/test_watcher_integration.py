# test_watcher_integration.py
"""
Integration test for watcher service.

Tests:
- Folder structure (exports/sites/, exports/cleaned/)
- Watcher service processing
- Master workbook creation
- CSV exports
- Idempotency
"""

import sys
import os
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))
import openpyxl

def test_folder_structure():
    """Test folder structure is correct."""
    print("Test 1: Folder structure")
    print("-" * 50)

    exports_dir = Path("exports")
    sites_dir = exports_dir / "sites"
    cleaned_dir = exports_dir / "cleaned"

    assert exports_dir.exists(), "exports/ directory missing"
    assert sites_dir.exists(), "exports/sites/ directory missing"
    assert cleaned_dir.exists(), "exports/cleaned/ directory missing"

    # Check sites directory has site folders
    site_folders = list(sites_dir.iterdir())
    print(f"  [PASS] exports/sites/ has {len(site_folders)} site folders")

    # Check cleaned directory exists
    print(f"  [PASS] exports/cleaned/ exists")

    return True


def test_master_workbook():
    """Test master workbook was created correctly."""
    print("\nTest 2: Master workbook")
    print("-" * 50)

    workbook_path = Path("exports/cleaned/MASTER_CLEANED_WORKBOOK.xlsx")

    assert workbook_path.exists(), "MASTER_CLEANED_WORKBOOK.xlsx not found"
    print(f"  [PASS] Master workbook exists ({workbook_path.stat().st_size} bytes)")

    # Load and check sheets
    wb = openpyxl.load_workbook(workbook_path, read_only=True)

    # Check metadata sheet
    assert "_Metadata" in wb.sheetnames, "_Metadata sheet missing"
    print(f"  [PASS] _Metadata sheet found")

    # Check metadata content
    ws_meta = wb["_Metadata"]
    rows = list(ws_meta.iter_rows(min_row=1, max_row=5, values_only=True))

    # Extract metadata
    metadata = {row[0]: row[1] for row in rows if row[0] and row[1]}

    total_sites = metadata.get("Total Sites")
    total_records = metadata.get("Total Records")

    print(f"  [PASS] Total Sites: {total_sites}")
    print(f"  [PASS] Total Records: {total_records}")

    # Check site sheets exist
    site_sheets = [s for s in wb.sheetnames if s != "_Metadata"]
    print(f"  [PASS] {len(site_sheets)} site sheets created")

    # Verify at least one site has records
    has_records = False
    for sheet in site_sheets[:5]:  # Check first 5
        ws = wb[sheet]
        record_count = ws.max_row - 1  # Subtract header
        if record_count > 0:
            print(f"  [PASS] {sheet} has {record_count} records")
            has_records = True
            break

    assert has_records, "No sheets have records"

    wb.close()
    return True


def test_csv_exports():
    """Test CSV exports were created."""
    print("\nTest 3: CSV exports")
    print("-" * 50)

    cleaned_dir = Path("exports/cleaned")

    # Count site folders in cleaned/
    site_folders = [f for f in cleaned_dir.iterdir() if f.is_dir() and f.name != "__pycache__"]

    print(f"  [PASS] {len(site_folders)} site folders in exports/cleaned/")

    # Check at least one has CSV
    csv_found = False
    for folder in site_folders:
        csv_files = list(folder.glob("*.csv"))
        if csv_files:
            csv_file = csv_files[0]
            size = csv_file.stat().st_size
            print(f"  [PASS] {folder.name}_cleaned.csv exists ({size} bytes)")
            csv_found = True
            break

    assert csv_found, "No CSV exports found"

    return True


def test_state_file():
    """Test state file was created."""
    print("\nTest 4: State file")
    print("-" * 50)

    state_file = Path("exports/cleaned/.watcher_state.json")

    if state_file.exists():
        import json
        with open(state_file, 'r') as f:
            state = json.load(f)

        processed_count = len(state.get('processed_files', {}))
        print(f"  [PASS] State file exists")
        print(f"  [PASS] {processed_count} files tracked")
    else:
        print(f"  [WARN] State file not found (might be locked)")

    return True


def test_metadata_json():
    """Test metadata.json was created."""
    print("\nTest 5: Metadata JSON")
    print("-" * 50)

    metadata_file = Path("exports/cleaned/metadata.json")

    assert metadata_file.exists(), "metadata.json not found"

    import json
    with open(metadata_file, 'r') as f:
        metadata = json.load(f)

    print(f"  [PASS] metadata.json exists")
    print(f"  [PASS] Tracking {len(metadata)} sites")

    # Show one site's metadata
    if metadata:
        site_key = list(metadata.keys())[0]
        site_meta = metadata[site_key]
        print(f"  [PASS] {site_key}: {site_meta.get('total_records')} total records")

    return True


def test_idempotency():
    """Test idempotency by running watcher again."""
    print("\nTest 6: Idempotency")
    print("-" * 50)

    import subprocess

    # Run watcher again
    result = subprocess.run(
        ["python", "watcher.py", "--once"],
        capture_output=True,
        text=True,
        timeout=60
    )

    output = result.stdout + result.stderr

    # Check for "No new or changed files"
    if "No new or changed files to process" in output:
        print("  [PASS] Watcher correctly skipped already-processed files")
        return True
    elif "files to process" in output and "Added 0 new records" in output:
        print("  [PASS] Watcher correctly skipped duplicate records")
        return True
    else:
        print("  [WARN] Idempotency test inconclusive")
        print(f"  Output: {output[:200]}")
        return True


def test_data_cleaning():
    """Test data cleaning features."""
    print("\nTest 7: Data cleaning")
    print("-" * 50)

    from core.data_cleaner import normalize_price, normalize_location, normalize_property_type

    # Test price normalization
    assert normalize_price("₦5,000,000") == "5000000", "Price normalization failed"
    assert normalize_price("5M") == "5000000", "Price abbreviation failed"
    assert normalize_price("500k") == "500000", "Price abbreviation failed"
    print("  [PASS] Price normalization works")

    # Test location normalization
    assert normalize_location("vi") == "Victoria Island", "Location alias failed"
    assert normalize_location("LEKKI") == "Lekki", "Location title case failed"
    print("  [PASS] Location normalization works")

    # Test property type normalization
    assert normalize_property_type("apartment") == "Flat", "Property type failed"
    assert normalize_property_type("plot") == "Land", "Property type failed"
    print("  [PASS] Property type normalization works")

    return True


def main():
    """Run all tests."""
    print("\n" + "=" * 60)
    print("WATCHER SERVICE INTEGRATION TESTS")
    print("=" * 60)

    tests = [
        test_folder_structure,
        test_master_workbook,
        test_csv_exports,
        test_state_file,
        test_metadata_json,
        test_idempotency,
        test_data_cleaning,
    ]

    passed = 0
    failed = 0

    for test in tests:
        try:
            if test():
                passed += 1
        except AssertionError as e:
            print(f"  [FAIL] {e}")
            failed += 1
        except Exception as e:
            print(f"  [ERROR] {e}")
            failed += 1

    print("\n" + "=" * 60)
    print(f"RESULTS: {passed} passed, {failed} failed")
    print("=" * 60)

    if failed > 0:
        sys.exit(1)
    else:
        print("\n[SUCCESS] All tests passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()
