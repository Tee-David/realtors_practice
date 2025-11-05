"""
Test master workbook file locking to verify multi-process safety
"""
import multiprocessing
import time
import logging
from pathlib import Path
from core.master_workbook import MasterWorkbookManager, FileLock, LOCK_FILE

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(process)d - %(levelname)s - %(message)s'
)

TEST_WORKBOOK = Path("exports/cleaned/TEST_MASTER_WORKBOOK.xlsx")

def test_file_lock():
    """Test the FileLock class"""
    print("\n=== Testing FileLock ===\n")

    lock = FileLock(LOCK_FILE, timeout=10)

    # Test acquire and release
    print("Test 1: Acquire and release lock")
    if lock.acquire():
        print("  [OK] Lock acquired")
        time.sleep(1)
        lock.release()
        print("  [OK] Lock released")
    else:
        print("  [FAIL] Could not acquire lock")
        return False

    # Test context manager
    print("\nTest 2: Context manager")
    try:
        with lock:
            print("  [OK] Lock acquired via context manager")
            time.sleep(1)
        print("  [OK] Lock released via context manager")
    except Exception as e:
        print(f"  [FAIL] Context manager error: {e}")
        return False

    print("\n[OK] FileLock tests passed")
    return True

def worker_process(worker_id: int, num_records: int):
    """Worker process that writes to the master workbook"""
    try:
        logging.info(f"Worker {worker_id} starting...")

        # Create fake records
        records = []
        for i in range(num_records):
            records.append({
                'hash': f'worker{worker_id}_record{i}',
                'title': f'Property {i} from worker {worker_id}',
                'price': 1000000 + (worker_id * 100000) + i,
                'location': 'Lagos',
                'property_type': 'Apartment',
                'bedrooms': 3,
                'source': f'worker_{worker_id}',
                'scrape_timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
                'listing_url': f'http://example.com/property_{worker_id}_{i}'
            })

        # Append to workbook
        manager = MasterWorkbookManager(TEST_WORKBOOK)
        site_key = f'test_site_{worker_id}'

        start_time = time.time()
        new_count = manager.append_to_site(site_key, records)
        elapsed = time.time() - start_time

        logging.info(f"Worker {worker_id} completed: {new_count} records added in {elapsed:.2f}s")
        return True

    except Exception as e:
        logging.error(f"Worker {worker_id} failed: {e}")
        return False

def test_concurrent_writes():
    """Test concurrent writes from multiple processes"""
    print("\n=== Testing Concurrent Writes ===\n")

    # Clean up test workbook if it exists
    if TEST_WORKBOOK.exists():
        TEST_WORKBOOK.unlink()
        print(f"Removed existing test workbook: {TEST_WORKBOOK}")

    # Create test workbook
    manager = MasterWorkbookManager(TEST_WORKBOOK)
    manager._ensure_workbook_exists()
    print(f"Created test workbook: {TEST_WORKBOOK}")

    # Launch multiple worker processes
    num_workers = 5
    records_per_worker = 10

    print(f"\nLaunching {num_workers} worker processes...")
    print(f"Each worker will write {records_per_worker} records")

    start_time = time.time()

    # Use multiprocessing to simulate batch scraping
    with multiprocessing.Pool(processes=num_workers) as pool:
        results = pool.starmap(
            worker_process,
            [(i, records_per_worker) for i in range(num_workers)]
        )

    elapsed = time.time() - start_time

    # Check results
    success_count = sum(1 for r in results if r)

    print(f"\n=== Results ===")
    print(f"Total time: {elapsed:.2f}s")
    print(f"Successful workers: {success_count}/{num_workers}")

    if success_count == num_workers:
        print("[OK] All workers completed successfully")
    else:
        print(f"[FAIL] {num_workers - success_count} workers failed")
        return False

    # Verify workbook integrity
    print("\nVerifying workbook integrity...")
    try:
        from openpyxl import load_workbook

        wb = load_workbook(TEST_WORKBOOK, read_only=True)
        print(f"  Sheets: {wb.sheetnames}")

        total_records = 0
        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith('_'):
                ws = wb[sheet_name]
                record_count = ws.max_row - 1  # Subtract header
                print(f"    {sheet_name}: {record_count} records")
                total_records += record_count

        wb.close()

        expected_records = num_workers * records_per_worker
        print(f"\n  Total records: {total_records}")
        print(f"  Expected: {expected_records}")

        if total_records == expected_records:
            print("[OK] All records written correctly")
            return True
        else:
            print(f"[FAIL] Record count mismatch")
            return False

    except Exception as e:
        print(f"[FAIL] Error verifying workbook: {e}")
        return False

def cleanup():
    """Clean up test files"""
    print("\nCleaning up test files...")

    if TEST_WORKBOOK.exists():
        TEST_WORKBOOK.unlink()
        print(f"  Removed {TEST_WORKBOOK}")

    if LOCK_FILE.exists():
        LOCK_FILE.unlink()
        print(f"  Removed {LOCK_FILE}")

if __name__ == "__main__":
    print("="*80)
    print("Master Workbook Locking Test")
    print("="*80)

    try:
        # Test 1: File locking
        if not test_file_lock():
            print("\n[FAIL] File lock tests failed")
            exit(1)

        # Test 2: Concurrent writes
        if not test_concurrent_writes():
            print("\n[FAIL] Concurrent write tests failed")
            exit(1)

        print("\n" + "="*80)
        print("[OK] ALL TESTS PASSED")
        print("="*80)

    except KeyboardInterrupt:
        print("\n\nTest interrupted by user")
    except Exception as e:
        print(f"\n[ERROR] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        cleanup()
